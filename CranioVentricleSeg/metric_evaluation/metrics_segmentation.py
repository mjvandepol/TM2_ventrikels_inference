# This file calculates Dice, NSD, volumes (RVE, AVE)
# Uses median (no CSP unless present) + summary per model (median + IQR + 95% CI)

import os
import numpy as np
import SimpleITK as sitk
import pandas as pd

from surface_distance import compute_surface_distances, compute_surface_dice_at_tolerance
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows


# =====================
# PATHS
# =====================

ground_truth_folder = "/data/scratch/r116411/ventricle_segmentation_train_test_t2/test/test_ground_truth"

model_folders = {
    "v1.0": "/data/scratch/r116411/ventricle_segmentation_train_test_t2/test/test_inference_v1.0",
    "v2.0": "/data/scratch/r116411/ventricle_segmentation_train_test_t2/test/test_inference_v2.0",
}

output_path = "/data/scratch/r116411/ventricle_segmentation_train_test_t2/test/segmentation_metrics_v2.xlsx"


# =====================
# LABELS
# =====================

label_names = {
    1: "RV",
    2: "3V",
    3: "4V",
    4: "CSP",
    5: "LV"
}

ordered_labels = ["RV", "3V", "4V", "CSP", "LV"]
csp_label = 4
nsd_tolerance_mm = 2.0


# =====================
# BASIC FUNCTIONS
# =====================

def load_image(path):
    img = sitk.ReadImage(path)
    arr = sitk.GetArrayFromImage(img)
    spacing = img.GetSpacing()[::-1]
    return arr, spacing


def dice_score(pred, gt):
    intersection = np.sum((pred == 1) & (gt == 1))
    size_sum = np.sum(pred == 1) + np.sum(gt == 1)

    if size_sum == 0:
        return np.nan

    return 2.0 * intersection / size_sum


def nsd_score(pred, gt, spacing):
    if np.sum(pred) == 0 and np.sum(gt) == 0:
        return np.nan

    distances = compute_surface_distances(
        gt.astype(bool),
        pred.astype(bool),
        spacing_mm=spacing
    )

    return compute_surface_dice_at_tolerance(distances, nsd_tolerance_mm)


# =====================
# VOLUMES
# =====================

def compute_volumes(arr, spacing):

    voxel_volume = spacing[0] * spacing[1] * spacing[2]

    volumes = {}

    for label, name in label_names.items():
        voxels = np.sum(arr == label)
        volumes[name] = (voxels * voxel_volume) / 1000  # ml

    total = sum(volumes.values())

    return volumes, total


def compute_rve(pred_vol, gt_vol):
    if gt_vol == 0:
        return np.nan
    return (pred_vol - gt_vol) / gt_vol


def compute_ave(pred_vol, gt_vol):
    return abs(pred_vol - gt_vol)


# =====================
# BOOTSTRAP CI (MEDIAN)
# =====================

def bootstrap_ci(data, n_boot=1000):

    data = np.array(data)
    data = data[~np.isnan(data)]

    if len(data) == 0:
        return np.nan, np.nan

    samples = np.random.choice(data, (n_boot, len(data)), replace=True)
    medians = np.median(samples, axis=1)

    return np.percentile(medians, 2.5), np.percentile(medians, 97.5)


# =====================
# BUILD DATAFRAME
# =====================

def build_dataframe(model_folder):

    rows = []
    gt_files = sorted(os.listdir(ground_truth_folder))

    for file in gt_files:

        print(f"\nProcessing case: {file}")

        gt_path = os.path.join(ground_truth_folder, file)
        pred_path = os.path.join(model_folder, file)

        if not os.path.exists(pred_path):
            print("  -> Prediction missing, skipping")
            continue

        gt, spacing = load_image(gt_path)
        pred, _ = load_image(pred_path)

        print("  -> Loaded images")

        gt_volumes, gt_total = compute_volumes(gt, spacing)
        pred_volumes, pred_total = compute_volumes(pred, spacing)

        row = {"case": file}

        dice_list = []
        nsd_list = []

        csp_in_gt = np.any(gt == csp_label)
        csp_in_pred = np.any(pred == csp_label)

        for label, name in label_names.items():

            gt_bin = (gt == label).astype(np.uint8)
            pred_bin = (pred == label).astype(np.uint8)

            d = dice_score(pred_bin, gt_bin)
            n = nsd_score(pred_bin, gt_bin, spacing)

            row[f"dice_{name}"] = d
            row[f"nsd_{name}"] = n

            row[f"GT_{name}"] = gt_volumes[name]
            row[f"Pred_{name}"] = pred_volumes[name]
            row[f"rve_{name}"] = compute_rve(pred_volumes[name], gt_volumes[name])
            row[f"ave_{name}"] = compute_ave(pred_volumes[name], gt_volumes[name])

            # CSP logic
            if name != "CSP":
                if not np.isnan(d):
                    dice_list.append(d)
                if not np.isnan(n):
                    nsd_list.append(n)
            else:
                if csp_in_gt and csp_in_pred:
                    if not np.isnan(d):
                        dice_list.append(d)
                    if not np.isnan(n):
                        nsd_list.append(n)

        row["dice_case_median"] = np.median(dice_list) if dice_list else np.nan
        row["nsd_case_median"] = np.median(nsd_list) if nsd_list else np.nan

        row["GT_total"] = gt_total
        row["Pred_total"] = pred_total
        row["rve_total"] = compute_rve(pred_total, gt_total)
        row["ave_total"] = compute_ave(pred_total, gt_total)

        print(f"  -> Done | Dice median: {row['dice_case_median']:.3f}")

        rows.append(row)

    return pd.DataFrame(rows)


# =====================
# SUMMARY (PER MODEL)
# =====================

def compute_summary(df):

    summary = {}

    for name in ["RV", "3V", "4V", "LV"]:

        vals = df[f"dice_{name}"].dropna()

        summary[f"{name}_median"] = vals.median()
        summary[f"{name}_q1"] = vals.quantile(0.25)
        summary[f"{name}_q3"] = vals.quantile(0.75)

        ci_low, ci_high = bootstrap_ci(vals)
        summary[f"{name}_ci_low"] = ci_low
        summary[f"{name}_ci_high"] = ci_high

    vals = df["dice_case_median"].dropna()

    summary["overall_median"] = vals.median()
    summary["overall_q1"] = vals.quantile(0.25)
    summary["overall_q3"] = vals.quantile(0.75)

    ci_low, ci_high = bootstrap_ci(vals)
    summary["overall_ci_low"] = ci_low
    summary["overall_ci_high"] = ci_high

    return summary


# =====================
# WRITE METRICS SHEET
# =====================

def write_sheet(ws, df):

    top_headers = [
        "Case",
        "RV","", "3V","", "4V","", "CSP","", "LV","",
        "Overall per case","",
        "Volumes","","",""
    ]

    sub_headers = [
        "Case",
        "Dice","NSD",
        "Dice","NSD",
        "Dice","NSD",
        "Dice","NSD",
        "Dice","NSD",
        "Dice","NSD",
        "GT","Pred","RVE","AVE"
    ]

    ws.append(top_headers)
    ws.append(sub_headers)

    merges = [(2,3),(4,5),(6,7),(8,9),(10,11),(12,13),(14,17)]
    for start, end in merges:
        ws.merge_cells(start_row=1, start_column=start, end_row=1, end_column=end)

    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)

    for _, r in df.iterrows():

        row = [
            r["case"],
            r.get("dice_RV"), r.get("nsd_RV"),
            r.get("dice_3V"), r.get("nsd_3V"),
            r.get("dice_4V"), r.get("nsd_4V"),
            r.get("dice_CSP"), r.get("nsd_CSP"),
            r.get("dice_LV"), r.get("nsd_LV"),
            r.get("dice_case_median"), r.get("nsd_case_median"),
            r.get("GT_total"), r.get("Pred_total"),
            r.get("rve_total"), r.get("ave_total"),
        ]

        ws.append(row)


# =====================
# VOLUME SHEET
# =====================

def build_volume_dataframe(df):

    rows = []

    for _, r in df.iterrows():

        row = {"case": r["case"]}

        for name in ordered_labels:
            row[f"GT_{name}"] = r[f"GT_{name}"]
            row[f"Pred_{name}"] = r[f"Pred_{name}"]
            row[f"RVE_{name}"] = r[f"rve_{name}"]
            row[f"AVE_{name}"] = r[f"ave_{name}"]

        row["GT_total"] = r["GT_total"]
        row["Pred_total"] = r["Pred_total"]
        row["RVE_total"] = r["rve_total"]
        row["AVE_total"] = r["ave_total"]

        rows.append(row)

    return pd.DataFrame(rows)


def write_volume_sheet(ws, df):
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)


# =====================
# MAIN
# =====================

wb = Workbook()
wb.remove(wb.active)

for model_name, model_folder in model_folders.items():

    print(f"\n=== Processing {model_name} ===")

    df = build_dataframe(model_folder)
    summary = compute_summary(df)

    ws1 = wb.create_sheet(title=f"{model_name}_metrics")
    write_sheet(ws1, df)

    ws1.append([])
    ws1.append(["SUMMARY (per model)"])
    for k, v in summary.items():
        ws1.append([k, v])

    df_vol = build_volume_dataframe(df)
    ws2 = wb.create_sheet(title=f"{model_name}_volumes")
    write_volume_sheet(ws2, df_vol)

wb.save(output_path)

print(f"\nExcel saved to: {output_path}")