# This file calculates Dice, NSD, volumes (RVE, AVE)
# Uses TOTAL Dice (volume-weighted) as overall metric + cluster bootstrap CI

import os
import numpy as np
import SimpleITK as sitk
import pandas as pd

from surface_distance import compute_surface_distances, compute_surface_dice_at_tolerance
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows


# ----------- Defining paths
ground_truth_folder = "/data/scratch/r116411/ventricle_segmentation_train_test_t2/test/test_ground_truth"

model_folders = {
    "v1.0": "/data/scratch/r116411/ventricle_segmentation_train_test_t2/test/test_inference_v1.0",
    "v2.0": "/data/scratch/r116411/ventricle_segmentation_train_test_t2/test/test_inference_v2.0",
}

output_path = "/data/scratch/r116411/ventricle_segmentation_train_test_t2/test/segmentation_metrics_v2_cluster_bootstrap.xlsx"


# ----------- Labels
# Label 6 is defined as all pixels with label >= 1 (union of all ventricles)

label_names = {
    1: "RV",
    2: "3V",
    3: "4V",
    4: "CSP",
    5: "LV",
    6: "TOTAL"
}

ordered_labels = ["RV", "3V", "4V", "CSP", "LV", "TOTAL"]
nsd_tolerance_MM = 2.0


# ----------- Basic functions
def load_image(path):
    img = sitk.ReadImage(path)
    arr = sitk.GetArrayFromImage(img)
    spacing = img.GetSpacing()[::-1]
    return arr, spacing


def dice_score(pred, gt):
    intersection = np.sum((pred == 1) & (gt == 1))
    size_sum = np.sum(pred == 1) + np.sum(gt == 1)

    if size_sum == 0:
        return np.nan

    return 2.0 * intersection / size_sum


def nsd_score(pred, gt, spacing):
    if np.sum(pred) == 0 and np.sum(gt) == 0:
        return np.nan

    distances = compute_surface_distances(
        gt.astype(bool),
        pred.astype(bool),
        spacing_mm=spacing
    )

    return compute_surface_dice_at_tolerance(distances, nsd_tolerance_MM)


# ----------- Calculate volume

def compute_volumes(arr, spacing):

    voxel_volume = spacing[0] * spacing[1] * spacing[2] #xyz
    volumes = {}

    for label, name in label_names.items():
        if label == 6:                              #classify all pixel with label > 1 as label 6
            voxels = np.sum(arr >= 1)
        else:
            voxels = np.sum(arr == label)

        volumes[name] = (voxels * voxel_volume) / 1000

    return volumes


def compute_rve(pred_vol, gt_vol):
    if gt_vol == 0:
        return np.nan
    return (pred_vol - gt_vol) / gt_vol


def compute_ave(pred_vol, gt_vol):
    return abs(pred_vol - gt_vol)


# ----------- Bootstrapping
# Use cluster boostrapping, resampling with replacement, to calculate bias and CI

def bootstrap_ci_cluster(df, value_col, patient_col="patient", n_boot=1000):

    patients = df[patient_col].unique()
    medians = []

    for _ in range(n_boot):

        sampled_patients = np.random.choice(patients, size=len(patients), replace=True)

        sampled_df = pd.concat([
            df[df[patient_col] == p] for p in sampled_patients
        ])

        medians.append(sampled_df[value_col].median())

    return np.percentile(medians, 2.5), np.percentile(medians, 97.5)


# ----------- Setup dataframe

def build_dataframe(model_folder):

    rows = []
    gt_files = sorted(os.listdir(ground_truth_folder))

    for file in gt_files:

        print(f"\nProcessing case: {file}")

        gt_path = os.path.join(ground_truth_folder, file)
        pred_path = os.path.join(model_folder, file)

        if not os.path.exists(pred_path):
            print("  -> Prediction missing, skipping")
            continue

        gt, spacing = load_image(gt_path)
        pred, _ = load_image(pred_path)

        print("  -> Loaded images")

        gt_volumes = compute_volumes(gt, spacing)
        pred_volumes = compute_volumes(pred, spacing)

        row = {"case": file}


        for label, name in label_names.items():

            if label == 6:
                gt_bin = (gt >= 1).astype(np.uint8)
                pred_bin = (pred >= 1).astype(np.uint8)
            else:
                gt_bin = (gt == label).astype(np.uint8)
                pred_bin = (pred == label).astype(np.uint8)

            d = dice_score(pred_bin, gt_bin)
            n = nsd_score(pred_bin, gt_bin, spacing)

            row[f"dice_{name}"] = d
            row[f"nsd_{name}"] = n

            row[f"GT_{name}"] = gt_volumes[name]
            row[f"Pred_{name}"] = pred_volumes[name]
            row[f"rve_{name}"] = compute_rve(pred_volumes[name], gt_volumes[name])
            row[f"ave_{name}"] = compute_ave(pred_volumes[name], gt_volumes[name])


        print(f"  -> Done | Dice total: {row['dice_TOTAL']:.3f}")

        rows.append(row)

    df = pd.DataFrame(rows)

    df["patient"] = df["case"].apply(lambda x: x.split("_")[0])

    return df

# ----------- Write volume sheet  

def write_volume_sheet(ws, df):

    top_headers = [
        "Case",
        "Right ventricle","","","",
        "3th ventricle","","","",
        "4th ventricle","","","",
        "CSP","","","",
        "Left ventricle","","","",
        "TOTAL",""
    ]

    sub_headers = [
        "Case",
        "GT","Pred","RVE","AVE",
        "GT","Pred","RVE","AVE",
        "GT","Pred","RVE","AVE",
        "GT","Pred","RVE","AVE",
        "GT","Pred","RVE","AVE",
        "GT","Pred"
    ]

    ws.append(top_headers)
    ws.append(sub_headers)

    merges = [(2,5),(6,9),(10,13),(14,17),(18,21),(22,23)]
    for start, end in merges:
        ws.merge_cells(start_row=1, start_column=start, end_row=1, end_column=end)

    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)

    for _, r in df.iterrows():
        ws.append([
            r["case"],
            r["GT_RV"], r["Pred_RV"], r["rve_RV"], r["ave_RV"],
            r["GT_3V"], r["Pred_3V"], r["rve_3V"], r["ave_3V"],
            r["GT_4V"], r["Pred_4V"], r["rve_4V"], r["ave_4V"],
            r["GT_CSP"], r["Pred_CSP"], r["rve_CSP"], r["ave_CSP"],
            r["GT_LV"], r["Pred_LV"], r["rve_LV"], r["ave_LV"],
            r["GT_total"], r["Pred_total"]
        ])

# ----------- Summary setup

def compute_summary(df):

    summary = {}

    for name in ["RV", "3V", "4V", "LV", "TOTAL"]:      #exclude CSP in summary

        vals = df[f"dice_{name}"].dropna()

        summary[f"{name}_median"] = vals.median()
        summary[f"{name}_q1"] = vals.quantile(0.25)
        summary[f"{name}_q3"] = vals.quantile(0.75)

        ci_low, ci_high = bootstrap_ci_cluster(df, f"dice_{name}")
        summary[f"{name}_ci_low"] = ci_low
        summary[f"{name}_ci_high"] = ci_high

    vals = df["dice_TOTAL"].dropna()

    summary["overall_median"] = vals.median()
    summary["overall_q1"] = vals.quantile(0.25)
    summary["overall_q3"] = vals.quantile(0.75)

    ci_low, ci_high = bootstrap_ci_cluster(df, "dice_TOTAL")
    summary["overall_ci_low"] = ci_low
    summary["overall_ci_high"] = ci_high

    return summary


# ----------- Write excel

wb = Workbook()
wb.remove(wb.active)

for model_name, model_folder in model_folders.items():

    print(f"\n=== Processing {model_name} ===")

    df = build_dataframe(model_folder)
    summary = compute_summary(df)

    ws1 = wb.create_sheet(title=f"{model_name}_metrics")
    metrics_cols = ["case"]

    for label in ordered_labels:
        metrics_cols.append(f"dice_{label}")

    for label in ordered_labels:
        metrics_cols.append(f"nsd_{label}")
    df_metrics = df[metrics_cols]

    for r in dataframe_to_rows(df_metrics, index=False, header=True):
        ws1.append(r)

    ws1.append([])
    ws1.append(["SUMMARY"])
    for k, v in summary.items():
        ws1.append([k, v])

    ws2 = wb.create_sheet(title=f"{model_name}_volumes")
    write_volume_sheet(ws2, df)

wb.save(output_path)

print(f"\nExcel saved to: {output_path}")
